import re
from collections import OrderedDict
from typing import List

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

class_mapper = dict(
    font=Font,
    fill=PatternFill,
    alignment=Alignment,
    border=Border,
)

styles = OrderedDict((
    ('danger', dict(
        font=dict(color='FF000000'),
        fill=dict(fill_type='solid', fgColor='FFFF0000'),
    )),
    ('red', dict(
        font=dict(color='FF9C0006'),
        fill=dict(fill_type='solid', fgColor='FFFFC7CE'),
    )),
    ('yellow', dict(
        font=dict(color='FF9C6500'),
        fill=dict(fill_type='solid', fgColor='FFFFEB9C'),
    )),
    ('green', dict(
        font=dict(color='FF006100'),
        fill=dict(fill_type='solid', fgColor='FFC6EFCE'),
    )),
    ('clear', dict(
        font=dict(color='FF000000'),
        fill=dict(fill_type='solid', fgColor='FFFFFFFF'),
    )),
    ('bold', dict(
        font=dict(bold=True),
    )),
    ('italic', dict(
        font=dict(italic=True),
    )),
    ('underline', dict(
        font=dict(underline="single"),
    )),
    ('center', dict(
        alignment=dict(horizontal='center'),
    )),
    ('right', dict(
        alignment=dict(horizontal='right'),
    )),
    ('left', dict(
        alignment=dict(horizontal='left'),
    )),
    ('middle', dict(
        alignment=dict(vertical='center'),
    )),
    ('top', dict(
        alignment=dict(vertical='top'),
    )),
    ('bottom', dict(
        alignment=dict(vertical='bottom'),
    )),
    ('wrap', dict(
        alignment=dict(wrap_text=True),
    )),
    ('celled', dict(
        border=dict(
            left=Side(style='thin', color='FF000000'),
            right=Side(style='thin', color='FF000000'),
            top=Side(style='thin', color='FF000000'),
            bottom=Side(style='thin', color='FF000000'),
        )
    )),
    ('comma', dict(
        number_format=numbers.FORMAT_NUMBER_COMMA_SEPARATED1
    )),
))


def apply_style(cell, style):
    group = {}
    for class_key in styles.keys():
        if re.search(r'\b{}\b'.format(class_key), style):
            class_style = styles[class_key]
            for style_key, style_value in class_style.items():
                if type(style_value) is dict:
                    group.setdefault(style_key, {}).update(style_value)
                else:
                    group[style_key] = style_value

    for key, kwargs in group.items():
        if key in class_mapper:
            setattr(cell, key, class_mapper[key](**kwargs))
        else:
            setattr(cell, key, kwargs)


class Column:
    def __init__(self, key, name, size, style='', formatter=None, stylish=None):
        self.key = key
        self.name = name
        self.size = size
        self.style = style
        self.formatter = formatter
        self.stylish = stylish

    def write_head(self, writer, index):
        writer.add_col(self.name, style=self.style + ' bold')
        writer.set_col_size(index + 1, self.size)

    def write_body(self, writer, row):
        value = row[self.key]
        if self.formatter:
            value = self.formatter(value, row)
        else:
            value = str(value)

        style = self.style
        if self.stylish:
            style = style + ' ' + self.stylish(value, row)

        writer.add_col(value, style=style)


class ExcelBytesWriter:
    def __init__(self, file_name=None):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.file_name = file_name

        self.row_pointer = 1
        self.col_pointer = 1

        self.excel = {}
        self.max_row = 0
        self.max_col = 0

    def set_col_size(self, col, size):
        self.ws.column_dimensions[get_column_letter(col)].width = size

    def add_row(self):
        self.row_pointer += 1
        self.col_pointer = 1

    def add_col(self, value='', span=1, style='', rowspan=1):
        while (self.row_pointer, self.col_pointer) in self.excel:
            self.col_pointer += 1

        if span > 1 or rowspan > 1:
            self.ws.merge_cells(
                start_row=self.row_pointer,
                start_column=self.col_pointer,
                end_row=self.row_pointer + rowspan - 1,
                end_column=self.col_pointer + span - 1
            )

        for row in range(self.row_pointer, self.row_pointer + rowspan):
            for col in range(self.col_pointer, self.col_pointer + span):
                cell = WriteOnlyCell(self.ws, value=value)
                style and apply_style(cell, style)
                self.excel[(row, col)] = cell
                self._keep_maximum_value(row, col)

        self.col_pointer += span

    def _keep_maximum_value(self, row, col):
        self.max_row = max(self.max_row, row)
        self.max_col = max(self.max_col, col)

    def render(self):
        for row in range(self.max_row):
            self.ws.append([self.excel.get((row + 1, col + 1)) for col in range(self.max_col)])

        return save_virtual_workbook(self.wb)

    def write_table(self, columns: List[Column], rows: List[dict]):
        self.add_row()
        for i, col in enumerate(columns):
            col.write_head(self, i)

        for row in rows:
            self.add_row()
            for col in columns:
                col.write_body(self, row)
